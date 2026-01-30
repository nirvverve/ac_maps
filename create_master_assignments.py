import pandas as pd
import numpy as np
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Disable truncation
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_colwidth', None)
pd.set_option('display.width', None)

print("=" * 80)
print("MASTER ACCOUNT ASSIGNMENT FILE CREATION")
print("=" * 80)

# ============================================================================
# STEP 1: Load Phoenix zip code to area mappings
# ============================================================================
print("\n" + "=" * 80)
print("STEP 1: Loading Phoenix zip code to area mappings")
print("=" * 80)

# Load the ZIP Code Detail sheet which has ALL zip codes with their area assignments
phoenix_zip_mapping = pd.read_excel(
    '/home/ubuntu/Phoenix_3Branch_Final_Consolidation.xlsx',
    sheet_name='ZIP Code Detail'
)

print(f"Phoenix ZIP mapping loaded: {phoenix_zip_mapping.shape}")
print(f"Unique areas: {phoenix_zip_mapping['ConsolidatedArea'].unique()}")

# Create a clean zip to area mapping dictionary
zip_to_area = {}
for _, row in phoenix_zip_mapping.iterrows():
    zip_code = str(row['ShippingPostalCode']).strip()
    area = row['ConsolidatedArea']
    source = row['Source']
    zip_to_area[zip_code] = {
        'area': area,
        'source': source
    }

print(f"\nTotal ZIP codes mapped: {len(zip_to_area)}")

# Load Tucson integration data to identify peripheral zips assigned to Phoenix East
tucson_integration = pd.read_excel(
    '/home/ubuntu/Phoenix_3Branch_Final_Consolidation.xlsx',
    sheet_name='Tucson Integration'
)
tucson_peripheral_zips = set(tucson_integration['ShippingPostalCode'].astype(str).str.strip())
print(f"Tucson peripheral ZIP codes assigned to Phoenix East: {len(tucson_peripheral_zips)}")
print(f"Zips: {sorted(tucson_peripheral_zips)}")

# ============================================================================
# STEP 2: Load and process Phoenix active accounts
# ============================================================================
print("\n" + "=" * 80)
print("STEP 2: Processing Phoenix active accounts")
print("=" * 80)

phoenix_accounts = pd.read_excel('/home/ubuntu/Uploads/Residential Data for Phoenix.xlsx')
print(f"Phoenix accounts loaded: {phoenix_accounts.shape}")

# Clean up zip codes
phoenix_accounts['ZIP_Clean'] = phoenix_accounts['ShippingPostalCode'].astype(str).str.strip().str[:5]

# Map to areas
def map_phoenix_account(row):
    zip_code = row['ZIP_Clean']
    if zip_code in zip_to_area:
        mapping = zip_to_area[zip_code]
        return pd.Series({
            'Branch_Assignment': f"Phoenix {mapping['area']}",
            'Area': mapping['area'],
            'Market': 'Phoenix',
            'Special_Flag': 'Tucson-Associated (Phoenix East)' if mapping['source'] == 'Tucson-Unassigned' else ''
        })
    else:
        return pd.Series({
            'Branch_Assignment': 'Unassigned',
            'Area': 'Unassigned',
            'Market': 'Phoenix',
            'Special_Flag': 'Unassigned ZIP Code'
        })

phoenix_assignments = phoenix_accounts.apply(map_phoenix_account, axis=1)
phoenix_master = pd.concat([phoenix_accounts, phoenix_assignments], axis=1)

print("\nPhoenix account assignment summary:")
print(phoenix_master['Branch_Assignment'].value_counts())
print("\nSpecial flags:")
print(phoenix_master['Special_Flag'].value_counts())

# ============================================================================
# STEP 3: Load and process Tucson active accounts
# ============================================================================
print("\n" + "=" * 80)
print("STEP 3: Processing Tucson active accounts")
print("=" * 80)

tucson_accounts = pd.read_csv('/home/ubuntu/Uploads/Tucson CG Active List.csv')
print(f"Tucson accounts loaded: {tucson_accounts.shape}")

# Clean up zip codes
tucson_accounts['ZIP_Clean'] = tucson_accounts['ShippingPostalCode'].astype(str).str.strip().str[:5]

# Load Tucson branch assignments
tucson_mapping = pd.read_csv('/home/ubuntu/tucson_account_mapping.csv')
print(f"Tucson mapping loaded: {tucson_mapping.shape}")

# Create a mapping from customer number to branch
customer_to_branch = {}
for _, row in tucson_mapping.iterrows():
    customer_num = row['Customer_Number__c']
    branch = row['Proposed_Branch']
    zip_code = str(row['ZIP_Clean']).strip()
    customer_to_branch[customer_num] = {
        'branch': branch,
        'zip': zip_code
    }

# Map Tucson accounts
def map_tucson_account(row):
    customer_num = row['Customer_Number__c']
    zip_code = row['ZIP_Clean']
    
    # Check if this zip is in the peripheral list assigned to Phoenix East
    if zip_code in tucson_peripheral_zips:
        return pd.Series({
            'Branch_Assignment': 'Phoenix East',
            'Area': 'East',
            'Market': 'Phoenix',
            'Special_Flag': 'Tucson-Associated (Phoenix East)'
        })
    
    # Otherwise, assign based on Tucson mapping
    if customer_num in customer_to_branch:
        mapping = customer_to_branch[customer_num]
        branch = mapping['branch']
        
        if pd.notna(branch) and branch != '':
            # Simplify branch name
            if 'Branch 1' in str(branch):
                area_name = 'Area 1 - East & North'
            elif 'Branch 2' in str(branch):
                area_name = 'Area 2 - West & Central'
            else:
                area_name = 'Unassigned'
                
            return pd.Series({
                'Branch_Assignment': f"Tucson {area_name}",
                'Area': area_name,
                'Market': 'Tucson',
                'Special_Flag': ''
            })
    
    # If not found, mark as unassigned
    return pd.Series({
        'Branch_Assignment': 'Tucson Unassigned',
        'Area': 'Unassigned',
        'Market': 'Tucson',
        'Special_Flag': 'No branch assignment found'
    })

tucson_assignments = tucson_accounts.apply(map_tucson_account, axis=1)
tucson_master = pd.concat([tucson_accounts, tucson_assignments], axis=1)

print("\nTucson account assignment summary:")
print(tucson_master['Branch_Assignment'].value_counts())
print("\nSpecial flags:")
print(tucson_master['Special_Flag'].value_counts())

# ============================================================================
# STEP 4: Create standardized master file
# ============================================================================
print("\n" + "=" * 80)
print("STEP 4: Creating standardized master account file")
print("=" * 80)

# Standardize Phoenix data
phoenix_standardized = pd.DataFrame({
    'Account_ID': phoenix_master['Display Name'],
    'Customer_Number': '',
    'Street_Address': phoenix_master['ShippingStreet'],
    'City': phoenix_master['ShippingCity'],
    'Zip_Code': phoenix_master['ZIP_Clean'],
    'Status': 'Active',
    'Market': phoenix_master['Market'],
    'Branch_Assignment': phoenix_master['Branch_Assignment'],
    'Area': phoenix_master['Area'],
    'Service_Contract': phoenix_master['Service Contract Description'],
    'Territory': phoenix_master['Territory'],
    'Route': phoenix_master['Route Name'],
    'Maintenance_Day': phoenix_master['Maintenance Plan Day of Week'],
    'Special_Flag': phoenix_master['Special_Flag'],
    'Email': phoenix_master['Invoice Bill To Email']
})

# Standardize Tucson data
tucson_standardized = pd.DataFrame({
    'Account_ID': tucson_master['Name'],
    'Customer_Number': tucson_master['Customer_Number__c'],
    'Street_Address': tucson_master['ShippingStreet'],
    'City': tucson_master['ShippingCity'],
    'Zip_Code': tucson_master['ZIP_Clean'],
    'Status': tucson_master['Status'],
    'Market': tucson_master['Market'],
    'Branch_Assignment': tucson_master['Branch_Assignment'],
    'Area': tucson_master['Area'],
    'Service_Contract': tucson_master['Name.1'],
    'Territory': tucson_master['Short Branch Name'],
    'Route': '',
    'Maintenance_Day': '',
    'Special_Flag': tucson_master['Special_Flag'],
    'Email': ''
})

# Combine both datasets
master_accounts = pd.concat([phoenix_standardized, tucson_standardized], ignore_index=True)

print(f"\nTotal accounts in master file: {len(master_accounts)}")
print(f"Phoenix accounts: {len(phoenix_standardized)}")
print(f"Tucson accounts: {len(tucson_standardized)}")

# ============================================================================
# STEP 5: Create summary statistics
# ============================================================================
print("\n" + "=" * 80)
print("STEP 5: Creating summary statistics")
print("=" * 80)

# Overall summary by branch assignment
summary_by_branch = master_accounts.groupby('Branch_Assignment').agg({
    'Account_ID': 'count',
    'Market': lambda x: x.mode()[0] if len(x.mode()) > 0 else ''
}).reset_index()
summary_by_branch.columns = ['Branch_Assignment', 'Account_Count', 'Primary_Market']
summary_by_branch = summary_by_branch.sort_values('Account_Count', ascending=False)

print("\nSummary by Branch Assignment:")
print(summary_by_branch)

# Summary by market and area
market_summary = master_accounts.groupby(['Market', 'Area']).agg({
    'Account_ID': 'count'
}).reset_index()
market_summary.columns = ['Market', 'Area', 'Account_Count']
market_summary = market_summary.sort_values(['Market', 'Account_Count'], ascending=[True, False])

print("\nSummary by Market and Area:")
print(market_summary)

# Special flag summary
special_flag_summary = master_accounts[master_accounts['Special_Flag'] != ''].groupby('Special_Flag').agg({
    'Account_ID': 'count'
}).reset_index()
special_flag_summary.columns = ['Special_Flag', 'Account_Count']

print("\nSpecial Flag Summary:")
print(special_flag_summary)

# ============================================================================
# STEP 6: Create Excel file with formatting
# ============================================================================
print("\n" + "=" * 80)
print("STEP 6: Creating formatted Excel file")
print("=" * 80)

wb = Workbook()

# Define styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
tucson_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
special_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet 1: All Accounts
ws1 = wb.active
ws1.title = 'All Accounts'

# Write headers
for col_idx, col_name in enumerate(master_accounts.columns, 1):
    cell = ws1.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Write data
for row_idx, row_data in enumerate(master_accounts.itertuples(index=False), 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        
        # Highlight Tucson-associated accounts
        if col_idx == master_accounts.columns.get_loc('Special_Flag') + 1:
            if 'Tucson-Associated' in str(value):
                cell.fill = special_fill

# Auto-adjust column widths
for column in ws1.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws1.column_dimensions[column_letter].width = adjusted_width

# Sheet 2: Phoenix Only
ws2 = wb.create_sheet('Phoenix Accounts')
phoenix_data = master_accounts[master_accounts['Market'] == 'Phoenix']
for col_idx, col_name in enumerate(phoenix_data.columns, 1):
    cell = ws2.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for row_idx, row_data in enumerate(phoenix_data.itertuples(index=False), 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# Sheet 3: Tucson Only
ws3 = wb.create_sheet('Tucson Accounts')
tucson_data = master_accounts[master_accounts['Market'] == 'Tucson']
for col_idx, col_name in enumerate(tucson_data.columns, 1):
    cell = ws3.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for row_idx, row_data in enumerate(tucson_data.itertuples(index=False), 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# Sheet 4: Summary Statistics
ws4 = wb.create_sheet('Summary Statistics')

# Branch assignment summary
ws4.cell(row=1, column=1, value='SUMMARY BY BRANCH ASSIGNMENT').font = Font(bold=True, size=14)
ws4.merge_cells('A1:C1')

row_offset = 3
for col_idx, col_name in enumerate(summary_by_branch.columns, 1):
    cell = ws4.cell(row=row_offset, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row_data in enumerate(summary_by_branch.itertuples(index=False), row_offset + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# Market summary
row_offset = row_offset + len(summary_by_branch) + 3
ws4.cell(row=row_offset, column=1, value='SUMMARY BY MARKET AND AREA').font = Font(bold=True, size=14)
ws4.merge_cells(f'A{row_offset}:C{row_offset}')

row_offset += 2
for col_idx, col_name in enumerate(market_summary.columns, 1):
    cell = ws4.cell(row=row_offset, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, row_data in enumerate(market_summary.itertuples(index=False), row_offset + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# Save the Excel file
excel_filename = '/home/ubuntu/Master_Account_Branch_Assignments.xlsx'
wb.save(excel_filename)
print(f"\nExcel file saved: {excel_filename}")

# ============================================================================
# STEP 7: Create map visualization data files
# ============================================================================
print("\n" + "=" * 80)
print("STEP 7: Creating map visualization data files")
print("=" * 80)

# Phoenix map data - aggregate by zip code
phoenix_map_data = master_accounts[master_accounts['Market'] == 'Phoenix'].groupby(['Zip_Code', 'Area', 'Branch_Assignment']).agg({
    'Account_ID': 'count',
    'Special_Flag': lambda x: (x != '').sum()  # Count of accounts with special flags
}).reset_index()
phoenix_map_data.columns = ['Zip_Code', 'Area', 'Branch_Assignment', 'Active_Accounts', 'Special_Flag_Count']
phoenix_map_data = phoenix_map_data.sort_values('Active_Accounts', ascending=False)

# Save as JSON
phoenix_json = phoenix_map_data.to_dict(orient='records')
with open('/home/ubuntu/Phoenix_Zip_Code_Map_Data.json', 'w') as f:
    json.dump(phoenix_json, f, indent=2)
print(f"Phoenix map data saved: /home/ubuntu/Phoenix_Zip_Code_Map_Data.json")
print(f"Total Phoenix ZIP codes: {len(phoenix_map_data)}")

# Also save as CSV
phoenix_map_data.to_csv('/home/ubuntu/Phoenix_Zip_Code_Map_Data.csv', index=False)
print(f"Phoenix map data CSV saved: /home/ubuntu/Phoenix_Zip_Code_Map_Data.csv")

# Tucson map data
tucson_map_data = master_accounts[master_accounts['Market'] == 'Tucson'].groupby(['Zip_Code', 'Area', 'Branch_Assignment']).agg({
    'Account_ID': 'count',
    'Special_Flag': lambda x: (x != '').sum()
}).reset_index()
tucson_map_data.columns = ['Zip_Code', 'Area', 'Branch_Assignment', 'Active_Accounts', 'Special_Flag_Count']
tucson_map_data = tucson_map_data.sort_values('Active_Accounts', ascending=False)

# Save as JSON
tucson_json = tucson_map_data.to_dict(orient='records')
with open('/home/ubuntu/Tucson_Zip_Code_Map_Data.json', 'w') as f:
    json.dump(tucson_json, f, indent=2)
print(f"Tucson map data saved: /home/ubuntu/Tucson_Zip_Code_Map_Data.json")
print(f"Total Tucson ZIP codes: {len(tucson_map_data)}")

# Also save as CSV
tucson_map_data.to_csv('/home/ubuntu/Tucson_Zip_Code_Map_Data.csv', index=False)
print(f"Tucson map data CSV saved: /home/ubuntu/Tucson_Zip_Code_Map_Data.csv")

# ============================================================================
# STEP 8: Create summary report
# ============================================================================
print("\n" + "=" * 80)
print("STEP 8: Creating summary report")
print("=" * 80)

report = f"""# MASTER ACCOUNT ASSIGNMENT REPORT
Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}

## Executive Summary

This report documents the comprehensive account assignment process for the swimming pool service 
company's branch reorganization initiative in Phoenix and Tucson markets.

### Total Accounts Processed
- **Total Accounts**: {len(master_accounts):,}
- **Phoenix Market**: {len(phoenix_standardized):,}
- **Tucson Market**: {len(tucson_standardized):,}

### Phoenix 3-Area Structure
{summary_by_branch[summary_by_branch['Primary_Market'] == 'Phoenix'].to_string(index=False)}

### Tucson 2-Area Structure
{summary_by_branch[summary_by_branch['Primary_Market'] == 'Tucson'].to_string(index=False)}

### Special Assignments
{special_flag_summary.to_string(index=False) if len(special_flag_summary) > 0 else 'No special assignments'}

## Assignment Process

### 1. Phoenix Assignments
- Loaded {len(phoenix_accounts)} active Phoenix accounts
- Matched accounts to branch areas based on ZIP code
- Used the Phoenix 3-Branch Consolidation analysis for ZIP code to area mappings
- Successfully assigned {len(phoenix_standardized[phoenix_standardized['Area'] != 'Unassigned'])} accounts
- {len(phoenix_standardized[phoenix_standardized['Area'] == 'Unassigned'])} unassigned accounts (missing ZIP codes)

### 2. Tucson Assignments
- Loaded {len(tucson_accounts)} active Tucson accounts
- Matched accounts using the Tucson branch decentralization plan
- Identified {len(tucson_peripheral_zips)} peripheral ZIP codes reassigned to Phoenix East
- Successfully assigned {len(tucson_standardized[tucson_standardized['Area'] != 'Unassigned'])} accounts

### 3. Special Cases
- **Tucson-Associated Accounts in Phoenix East**: {len(master_accounts[master_accounts['Special_Flag'].str.contains('Tucson-Associated', na=False)])} accounts
  - These are accounts in peripheral areas (Casa Grande, Maricopa, San Tan Valley, etc.) 
  - Assigned to Phoenix East Area for operational purposes
  - Flagged as Tucson-associated for tracking

## Deliverables

### 1. Master Account Assignment File
**File**: `Master_Account_Branch_Assignments.xlsx`

Contains four sheets:
- **All Accounts**: Complete list of all {len(master_accounts)} accounts
- **Phoenix Accounts**: {len(phoenix_data)} Phoenix market accounts
- **Tucson Accounts**: {len(tucson_data)} Tucson market accounts
- **Summary Statistics**: Aggregated statistics by branch and market

### 2. Map Visualization Data
**Phoenix Data**: 
- `Phoenix_Zip_Code_Map_Data.json` ({len(phoenix_map_data)} ZIP codes)
- `Phoenix_Zip_Code_Map_Data.csv`

**Tucson Data**:
- `Tucson_Zip_Code_Map_Data.json` ({len(tucson_map_data)} ZIP codes)
- `Tucson_Zip_Code_Map_Data.csv`

These files contain:
- ZIP code
- Assigned area
- Number of active accounts
- Special flag counts

## Data Quality Notes

### Missing Data
- {len(master_accounts[master_accounts['Email'] == ''])} accounts missing email addresses
- {len(master_accounts[master_accounts['Route'] == ''])} accounts missing route assignments
- {len(master_accounts[master_accounts['Maintenance_Day'] == ''])} accounts missing maintenance day

### Unassigned Accounts
- {len(master_accounts[master_accounts['Branch_Assignment'].str.contains('Unassigned', na=False)])} accounts without proper branch assignments
- These accounts require manual review and assignment

## Recommendations

1. **Review Unassigned Accounts**: Investigate accounts with missing ZIP codes or branch assignments
2. **Validate Special Assignments**: Review Tucson-associated accounts assigned to Phoenix East
3. **Update Route Information**: Many accounts are missing route assignments
4. **Email Collection**: Improve email address collection for customer communication

## Next Steps

1. Review and validate the master assignment file
2. Use map visualization data to create geographic territory maps
3. Communicate new assignments to field teams
4. Update CRM systems with new branch assignments
5. Monitor retention rates by new branch structure
"""

with open('/home/ubuntu/Master_Account_Assignment_Report.md', 'w') as f:
    f.write(report)
print(f"\nReport saved: /home/ubuntu/Master_Account_Assignment_Report.md")

print("\n" + "=" * 80)
print("MASTER ACCOUNT ASSIGNMENT PROCESS COMPLETE!")
print("=" * 80)
print("\nFiles created:")
print("1. Master_Account_Branch_Assignments.xlsx")
print("2. Phoenix_Zip_Code_Map_Data.json")
print("3. Phoenix_Zip_Code_Map_Data.csv")
print("4. Tucson_Zip_Code_Map_Data.json")
print("5. Tucson_Zip_Code_Map_Data.csv")
print("6. Master_Account_Assignment_Report.md")
print("\nAll files saved to /home/ubuntu/")
